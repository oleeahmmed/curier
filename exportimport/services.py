"""
Services for bag management operations including PDF generation.
"""
from io import BytesIO
from django.core.files.base import ContentFile
from django.utils import timezone
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.units import mm, inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from django.db import transaction


class ManifestPDFGenerator:
    """
    Generate PDF export for manifest using ReportLab.
    Follows the format specified in manifast.html template.
    """
    
    def __init__(self, manifest):
        """
        Initialize generator with manifest instance.
        
        Args:
            manifest: Manifest instance to generate PDF for
        """
        self.manifest = manifest
        self.buffer = BytesIO()
        # Use landscape orientation
        from reportlab.lib.pagesizes import landscape
        self.page_width, self.page_height = landscape(A4)
    
    def generate(self):
        """
        Generate the PDF document.
        
        Returns:
            bytes: PDF content
        """
        from reportlab.lib.pagesizes import landscape
        
        # Create the PDF document with landscape orientation
        doc = SimpleDocTemplate(
            self.buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Build story (content)
        story = []
        
        # Add header
        story.extend(self._build_header())
        story.append(Spacer(1, 15))
        
        # Add meta section
        story.extend(self._build_meta_section())
        story.append(Spacer(1, 15))
        
        # Add shipments table
        story.append(self._build_shipments_table())
        story.append(Spacer(1, 10))
        
        # Add footer
        story.extend(self._build_footer())
        
        # Build the PDF
        doc.build(story)
        
        # Get the PDF content
        pdf_content = self.buffer.getvalue()
        self.buffer.close()
        
        return pdf_content
    
    def _build_header(self):
        """Build header section with company name and title."""
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'ManifestTitle',
            parent=styles['Normal'],
            fontSize=16,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            spaceAfter=10
        )
        
        content = []
        content.append(Paragraph("Fast Line Express", title_style))
        
        return content
    
    def _build_meta_section(self):
        """Build meta information section."""
        styles = getSampleStyleSheet()
        
        meta_style = ParagraphStyle(
            'MetaInfo',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_LEFT,
            spaceAfter=3
        )
        
        content = []
        
        # Format departure date
        departure_date_str = self.manifest.departure_date.strftime('%d/%m/%Y') if self.manifest.departure_date else 'N/A'
        
        # Create meta information table for better layout
        meta_data = [
            [
                Paragraph("<b>Agent Name:</b> FAST LINE (DAC)", meta_style),
                Paragraph(f"<b>Flight No.:</b> {self.manifest.flight_number or 'N/A'}", meta_style)
            ],
            [
                Paragraph(f"<b>MAWB:</b> {self.manifest.mawb_number or 'N/A'}", meta_style),
                Paragraph(f"<b>Flight Date:</b> {departure_date_str}", meta_style)
            ]
        ]
        
        meta_table = Table(meta_data, colWidths=[270, 270])
        meta_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        content.append(meta_table)
        
        return content
    
    def _build_shipments_table(self):
        """Build shipments table with proper text wrapping and no borders."""
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        
        # Create paragraph styles for wrapping text
        cell_style_left = ParagraphStyle(
            'CellLeft',
            fontName='Helvetica',
            fontSize=7,
            alignment=TA_LEFT,
            leading=9,
            wordWrap='CJK'
        )
        
        cell_style_center = ParagraphStyle(
            'CellCenter',
            fontName='Helvetica',
            fontSize=7,
            alignment=TA_CENTER,
            leading=9,
            wordWrap='CJK'
        )
        
        # Header style - no wrapping for headers
        header_style = ParagraphStyle(
            'HeaderStyle',
            fontName='Helvetica-Bold',
            fontSize=8,
            alignment=TA_CENTER,
            leading=10,
            wordWrap=None  # Prevent header wrapping
        )
        
        # Table headers - single line, no wrapping
        headers = [
            Paragraph('No.', header_style),
            Paragraph('AWB&nbsp;No', header_style),  # Non-breaking space
            Paragraph('Shipper', header_style),
            Paragraph('Consignee', header_style),
            Paragraph('Origin', header_style),
            Paragraph('Dest', header_style),
            Paragraph('PCS', header_style),
            Paragraph('Weight', header_style),
            Paragraph('Description', header_style),
            Paragraph('Curr', header_style),  # Shortened to prevent wrapping
            Paragraph('Value', header_style),
            Paragraph('Code', header_style),
            Paragraph('Remark', header_style),
            Paragraph('COD', header_style),
            Paragraph('Bag&nbsp;no', header_style)  # Non-breaking space
        ]
        
        # Build table data
        table_data = [headers]
        
        # Iterate through all bags and shipments
        shipment_number = 1
        for bag in self.manifest.bags.all():
            for shipment in bag.shipment.all():
                # Format consignee with full details (name, address, phone, city)
                consignee_lines = []
                if shipment.recipient_name:
                    consignee_lines.append(shipment.recipient_name)
                if shipment.recipient_address:
                    consignee_lines.append(shipment.recipient_address)
                if shipment.recipient_phone:
                    consignee_lines.append(f"Tel: {shipment.recipient_phone}")
                if shipment.recipient_country:
                    consignee_lines.append(shipment.recipient_country)
                
                consignee_text = '<br/>'.join(consignee_lines) if consignee_lines else ''
                
                # Format weight - empty if not available
                weight_str = f"{shipment.weight_estimated}" if shipment.weight_estimated else ""
                
                # Origin - shipper country (empty if not available)
                origin = shipment.shipper_country[:3].upper() if shipment.shipper_country else ""
                
                # Destination - recipient country (empty if not available)
                dest = shipment.recipient_country[:3].upper() if shipment.recipient_country else ""
                
                # Currency - empty if not available
                currency = shipment.declared_currency if shipment.declared_currency else ""
                
                # Value - empty if not available
                value = f"{shipment.declared_value}" if shipment.declared_value else ""
                
                # Code - "COD" only if is_cod is True, otherwise empty
                code = "COD" if shipment.is_cod else ""
                
                # COD amount - only show if COD and amount exists, otherwise empty
                cod_value = ""
                if shipment.is_cod and shipment.cod_amount:
                    cod_value = f"{shipment.cod_amount}"
                
                row = [
                    Paragraph(str(shipment_number), cell_style_center),
                    Paragraph(shipment.awb_number if shipment.awb_number else '', cell_style_left),
                    Paragraph(shipment.shipper_name if shipment.shipper_name else '', cell_style_left),
                    Paragraph(consignee_text, cell_style_left),
                    Paragraph(origin, cell_style_center),
                    Paragraph(dest, cell_style_center),
                    Paragraph('1', cell_style_center),  # PCS - default to 1 piece
                    Paragraph(weight_str, cell_style_center),
                    Paragraph(shipment.contents if shipment.contents else '', cell_style_left),
                    Paragraph(currency, cell_style_center),
                    Paragraph(value, cell_style_center),
                    Paragraph(code, cell_style_center),
                    Paragraph('', cell_style_center),  # Remark - empty
                    Paragraph(cod_value, cell_style_center),
                    Paragraph(bag.bag_number, cell_style_center)
                ]
                
                table_data.append(row)
                shipment_number += 1
        
        # Add individual shipments (not in bags)
        for shipment in self.manifest.shipments.all():
            # Format consignee with full details
            consignee_lines = []
            if shipment.recipient_name:
                consignee_lines.append(shipment.recipient_name)
            if shipment.recipient_address:
                consignee_lines.append(shipment.recipient_address)
            if shipment.recipient_phone:
                consignee_lines.append(f"Tel: {shipment.recipient_phone}")
            if shipment.recipient_country:
                consignee_lines.append(shipment.recipient_country)
            
            consignee_text = '<br/>'.join(consignee_lines) if consignee_lines else ''
            
            # Format weight - empty if not available
            weight_str = f"{shipment.weight_estimated}" if shipment.weight_estimated else ""
            
            # Origin - shipper country (empty if not available)
            origin = shipment.shipper_country[:3].upper() if shipment.shipper_country else ""
            
            # Destination - recipient country (empty if not available)
            dest = shipment.recipient_country[:3].upper() if shipment.recipient_country else ""
            
            # Currency - empty if not available
            currency = shipment.declared_currency if shipment.declared_currency else ""
            
            # Value - empty if not available
            value = f"{shipment.declared_value}" if shipment.declared_value else ""
            
            # Code - "COD" only if is_cod is True, otherwise empty
            code = "COD" if shipment.is_cod else ""
            
            # COD amount - only show if COD and amount exists, otherwise empty
            cod_value = ""
            if shipment.is_cod and shipment.cod_amount:
                cod_value = f"{shipment.cod_amount}"
            
            row = [
                Paragraph(str(shipment_number), cell_style_center),
                Paragraph(shipment.awb_number if shipment.awb_number else '', cell_style_left),
                Paragraph(shipment.shipper_name if shipment.shipper_name else '', cell_style_left),
                Paragraph(consignee_text, cell_style_left),
                Paragraph(origin, cell_style_center),
                Paragraph(dest, cell_style_center),
                Paragraph('1', cell_style_center),  # PCS - default to 1 piece
                Paragraph(weight_str, cell_style_center),
                Paragraph(shipment.contents if shipment.contents else '', cell_style_left),
                Paragraph(currency, cell_style_center),
                Paragraph(value, cell_style_center),
                Paragraph(code, cell_style_center),
                Paragraph('', cell_style_center),  # Remark - empty
                Paragraph(cod_value, cell_style_center),
                Paragraph('INDIVIDUAL', cell_style_center)  # Mark as individual shipment
            ]
            
            table_data.append(row)
            shipment_number += 1
        
        # Column widths (adjusted for landscape A4) - 15 columns now
        col_widths = [25, 60, 70, 100, 35, 35, 30, 40, 80, 40, 40, 35, 45, 35, 50]
        
        # Create table
        table = Table(table_data, colWidths=col_widths)
        
        # Table styling - NO BORDERS, only header background
        table.setStyle(TableStyle([
            # Headers - only background color, no borders
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            
            # Padding for all cells
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        return table
    
    def _build_footer(self):
        """Build footer with totals."""
        styles = getSampleStyleSheet()
        
        footer_style = ParagraphStyle(
            'FooterText',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold',
            alignment=TA_LEFT,
            spaceAfter=3
        )
        
        content = []
        
        # Calculate totals
        total_shipments = self.manifest.total_parcels
        total_pcs = total_shipments  # Assuming 1 piece per shipment
        total_weight = self.manifest.total_weight
        
        # Create footer text
        footer_text = f"TOTAL SHIPMENTS: {total_shipments} &nbsp;&nbsp;&nbsp; PCS: {total_pcs} &nbsp;&nbsp;&nbsp; WEIGHT: {total_weight} KGS"
        
        content.append(Paragraph(footer_text, footer_style))
        
        return content


class ManifestExcelGenerator:
    """
    Generate Excel export for manifest using openpyxl.
    Same structure as PDF but in Excel format.
    """
    
    def __init__(self, manifest):
        """
        Initialize generator with manifest instance.
        
        Args:
            manifest: Manifest instance to generate Excel for
        """
        self.manifest = manifest
        self.workbook = None
        self.worksheet = None
    
    def generate(self):
        """
        Generate the Excel document.
        
        Returns:
            bytes: Excel content
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        
        # Create workbook and worksheet
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Manifest"
        
        # Current row tracker
        self.current_row = 1
        
        # Build Excel content
        self._write_header()
        self._write_meta_info()
        self._write_column_headers()
        self._write_shipment_rows()
        self._write_totals_row()
        self._apply_formatting()
        
        # Save to BytesIO
        buffer = BytesIO()
        self.workbook.save(buffer)
        excel_content = buffer.getvalue()
        buffer.close()
        
        return excel_content
    
    def _write_header(self):
        """Write header section."""
        from openpyxl.styles import Font, Alignment
        
        # Write company name in first row
        self.worksheet.cell(row=self.current_row, column=1, value="Fast Line Express")
        self.worksheet.cell(row=self.current_row, column=1).font = Font(size=16, bold=True)
        self.worksheet.cell(row=self.current_row, column=1).alignment = Alignment(horizontal='center')
        self.worksheet.merge_cells(start_row=self.current_row, start_column=1, end_row=self.current_row, end_column=15)
        self.current_row += 1
        
        # Add blank row
        self.current_row += 1
    
    def _write_meta_info(self):
        """Write meta information."""
        from openpyxl.styles import Font
        
        # Format departure date
        departure_date_str = self.manifest.departure_date.strftime('%d/%m/%Y') if self.manifest.departure_date else 'N/A'
        
        # Write Agent Name and Flight No
        self.worksheet.cell(row=self.current_row, column=1, value="Agent Name:")
        self.worksheet.cell(row=self.current_row, column=1).font = Font(bold=True)
        self.worksheet.cell(row=self.current_row, column=2, value="FAST LINE (DAC)")
        
        self.worksheet.cell(row=self.current_row, column=4, value="Flight No.:")
        self.worksheet.cell(row=self.current_row, column=4).font = Font(bold=True)
        self.worksheet.cell(row=self.current_row, column=5, value=self.manifest.flight_number or 'N/A')
        self.current_row += 1
        
        # Write MAWB and Flight Date
        self.worksheet.cell(row=self.current_row, column=1, value="MAWB:")
        self.worksheet.cell(row=self.current_row, column=1).font = Font(bold=True)
        self.worksheet.cell(row=self.current_row, column=2, value=self.manifest.mawb_number or 'N/A')
        
        self.worksheet.cell(row=self.current_row, column=4, value="Flight Date:")
        self.worksheet.cell(row=self.current_row, column=4).font = Font(bold=True)
        self.worksheet.cell(row=self.current_row, column=5, value=departure_date_str)
        self.current_row += 1
        
        # Add blank row
        self.current_row += 1
    
    def _write_column_headers(self):
        """Write table column headers."""
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        # Column headers - matching PDF format
        headers = [
            'No.', 'AWB No', 'Shipper', 'Consignee', 'Origin', 'Dest', 'PCS',
            'Weight', 'Description', 'Currency', 'Value', 'Code', 'Remark', 'COD', 'Bag no'
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=self.current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        self.current_row += 1
    
    def _write_shipment_rows(self):
        """Write all shipment data rows."""
        from openpyxl.styles import Alignment
        
        # Iterate through all bags and shipments
        shipment_number = 1
        for bag in self.manifest.bags.all():
            for shipment in bag.shipment.all():
                # Consignee - NAME ONLY for Excel
                consignee_name = shipment.recipient_name or 'N/A'
                
                # Format weight
                weight_str = f"{shipment.weight_estimated}" if shipment.weight_estimated else "0"
                
                # Origin - shipper country
                origin = shipment.shipper_country[:3].upper() if shipment.shipper_country else "N/A"
                
                # Destination - recipient country
                dest = shipment.recipient_country[:3].upper() if shipment.recipient_country else "N/A"
                
                # Currency
                currency = shipment.declared_currency or "USD"
                
                # Value
                value = f"{shipment.declared_value}" if shipment.declared_value else "0.00"
                
                # Code - COD if applicable
                code = "COD" if shipment.is_cod else ""
                
                # COD amount or indicator
                cod_value = ""
                if shipment.is_cod and shipment.cod_amount:
                    cod_value = f"{shipment.cod_amount}"
                elif shipment.is_cod:
                    cod_value = "COD"
                
                # Write row data
                row_data = [
                    shipment_number,
                    shipment.awb_number or 'N/A',
                    shipment.shipper_name or 'N/A',
                    consignee_name,  # NAME ONLY
                    origin,
                    dest,
                    1,  # PCS - default to 1 piece
                    weight_str,
                    shipment.contents or 'N/A',
                    currency,
                    value,
                    code,
                    '',  # Remark - empty for now
                    cod_value,
                    bag.bag_number
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    cell = self.worksheet.cell(row=self.current_row, column=col_idx, value=value)
                    
                    # Set alignment based on column
                    if col_idx == 1:  # No. column
                        cell.alignment = Alignment(horizontal='center', vertical='top')
                    elif col_idx in [2, 3, 4]:  # AWB, Shipper, Consignee
                        cell.alignment = Alignment(horizontal='left', vertical='top')
                    else:  # Rest
                        cell.alignment = Alignment(horizontal='center', vertical='top')
                
                self.current_row += 1
                shipment_number += 1
        
        # Add individual shipments (not in bags)
        for shipment in self.manifest.shipments.all():
            # Consignee - NAME ONLY for Excel
            consignee_name = shipment.recipient_name or 'N/A'
            
            # Format weight
            weight_str = f"{shipment.weight_estimated}" if shipment.weight_estimated else "0"
            
            # Origin - shipper country
            origin = shipment.shipper_country[:3].upper() if shipment.shipper_country else "N/A"
            
            # Destination - recipient country
            dest = shipment.recipient_country[:3].upper() if shipment.recipient_country else "N/A"
            
            # Currency
            currency = shipment.declared_currency or "USD"
            
            # Value
            value = f"{shipment.declared_value}" if shipment.declared_value else "0.00"
            
            # Code - COD if applicable
            code = "COD" if shipment.is_cod else ""
            
            # COD amount or indicator
            cod_value = ""
            if shipment.is_cod and shipment.cod_amount:
                cod_value = f"{shipment.cod_amount}"
            elif shipment.is_cod:
                cod_value = "COD"
            
            # Write row data
            row_data = [
                shipment_number,
                shipment.awb_number or 'N/A',
                shipment.shipper_name or 'N/A',
                consignee_name,  # NAME ONLY
                origin,
                dest,
                1,  # PCS - default to 1 piece
                weight_str,
                shipment.contents or 'N/A',
                currency,
                value,
                code,
                '',  # Remark - empty for now
                cod_value,
                'INDIVIDUAL'  # Mark as individual shipment
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx, value=value)
                
                # Set alignment based on column
                if col_idx == 1:  # No. column
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                elif col_idx in [2, 3, 4]:  # AWB, Shipper, Consignee
                    cell.alignment = Alignment(horizontal='left', vertical='top')
                else:  # Rest
                    cell.alignment = Alignment(horizontal='center', vertical='top')
            
            self.current_row += 1
            shipment_number += 1
    
    def _write_totals_row(self):
        """Write totals row at bottom."""
        from openpyxl.styles import Font, Alignment
        
        # Add blank row before totals
        self.current_row += 1
        
        # Calculate totals
        total_shipments = self.manifest.total_parcels
        total_pcs = total_shipments  # Assuming 1 piece per shipment
        total_weight = self.manifest.total_weight
        
        # Write totals text
        totals_text = f"TOTAL SHIPMENTS: {total_shipments}     PCS: {total_pcs}     WEIGHT: {total_weight} KGS"
        self.worksheet.cell(row=self.current_row, column=1, value=totals_text)
        self.worksheet.cell(row=self.current_row, column=1).font = Font(bold=True, size=10)
        self.worksheet.cell(row=self.current_row, column=1).alignment = Alignment(horizontal='left')
        self.worksheet.merge_cells(start_row=self.current_row, start_column=1, end_row=self.current_row, end_column=15)
    
    def _apply_formatting(self):
        """Apply cell formatting and styling."""
        from openpyxl.styles import Border, Side
        from openpyxl.utils import get_column_letter
        
        # Set column widths - 15 columns now
        column_widths = {
            1: 6,   # No.
            2: 15,  # AWB No
            3: 18,  # Shipper
            4: 30,  # Consignee (wider for full details)
            5: 8,   # Origin
            6: 8,   # Dest
            7: 6,   # PCS
            8: 10,  # Weight
            9: 20,  # Description
            10: 10, # Currency
            11: 12, # Value
            12: 8,  # Code
            13: 12, # Remark
            14: 10, # COD
            15: 15  # Bag no
        }
        
        for col_idx, width in column_widths.items():
            column_letter = get_column_letter(col_idx)
            self.worksheet.column_dimensions[column_letter].width = width
        
        # Add borders to data table (from column headers to last shipment row)
        # Find the header row (should be row 7 based on structure)
        header_row = 7
        last_data_row = self.current_row - 2  # Exclude blank row and totals row
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to all cells in the table - 15 columns now
        for row in range(header_row, last_data_row + 1):
            for col in range(1, 16):  # 15 columns
                self.worksheet.cell(row=row, column=col).border = thin_border



class ManifestFinalizationService:
    """
    Handle the complete manifest finalization workflow.
    Encapsulates all operations that must happen atomically.
    """
    
    def __init__(self, manifest, user):
        """
        Initialize service.
        
        Args:
            manifest: Manifest instance to finalize
            user: User performing the finalization
        """
        self.manifest = manifest
        self.user = user
    
    def _validate_can_finalize(self):
        """
        Validate manifest is in correct state for finalization.
        
        Raises:
            ValidationError: If manifest cannot be finalized
        """
        from django.core.exceptions import ValidationError
        
        if self.manifest.status != 'DRAFT':
            raise ValidationError("Only DRAFT manifests can be finalized")
        
        # Check if manifest has at least one bag OR one individual shipment
        if self.manifest.bags.count() == 0 and self.manifest.shipments.count() == 0:
            raise ValidationError("Manifest must have at least one bag or shipment")
    
    def _update_manifest_status(self):
        """Update manifest to FINALIZED status."""
        self.manifest.status = 'FINALIZED'
        self.manifest.finalized_by = self.user
        self.manifest.finalized_at = timezone.now()
        self.manifest.save()
    
    def _update_bags(self):
        """Update all bags to IN_MANIFEST status."""
        for bag in self.manifest.bags.all():
            bag.status = 'IN_MANIFEST'
            bag.save()
    
    def _update_shipments(self):
        """Update all shipments to IN_EXPORT_MANIFEST status."""
        # Update shipments in bags
        for bag in self.manifest.bags.all():
            for shipment in bag.shipment.all():
                shipment.current_status = 'IN_EXPORT_MANIFEST'
                shipment.save()
        
        # Update individual shipments
        for shipment in self.manifest.shipments.all():
            shipment.current_status = 'IN_EXPORT_MANIFEST'
            shipment.save()
    
    def _create_tracking_events(self):
        """Create tracking events for all shipments."""
        from .models import TrackingEvent
        
        # Create events for shipments in bags
        for bag in self.manifest.bags.all():
            for shipment in bag.shipment.all():
                TrackingEvent.objects.create(
                    shipment=shipment,
                    status='IN_EXPORT_MANIFEST',
                    description=f"Added to manifest {self.manifest.manifest_number} (via bag {bag.bag_number})",
                    location='Bangladesh Warehouse',
                    updated_by=self.user
                )
        
        # Create events for individual shipments
        for shipment in self.manifest.shipments.all():
            TrackingEvent.objects.create(
                shipment=shipment,
                status='IN_EXPORT_MANIFEST',
                description=f"Added to manifest {self.manifest.manifest_number}",
                location='Bangladesh Warehouse',
                updated_by=self.user
            )
    
    def _generate_exports(self):
        """
        Generate PDF and Excel exports.
        
        Returns:
            tuple: (pdf_bytes, excel_bytes)
        """
        # Generate PDF
        pdf_generator = ManifestPDFGenerator(self.manifest)
        pdf_bytes = pdf_generator.generate()
        
        # Generate Excel
        excel_generator = ManifestExcelGenerator(self.manifest)
        excel_bytes = excel_generator.generate()
        
        return (pdf_bytes, excel_bytes)
    
    def _store_exports(self, pdf_content, excel_content):
        """
        Store exports in ManifestExport model.
        Delete and recreate if manifest was re-finalized.
        
        Args:
            pdf_content: PDF bytes
            excel_content: Excel bytes
        """
        from .models import ManifestExport
        
        # Check if export already exists (manifest was re-finalized)
        try:
            existing_export = ManifestExport.objects.get(manifest=self.manifest)
            # Delete the existing export completely
            existing_export.delete()
        except ManifestExport.DoesNotExist:
            pass
        
        # Create new ManifestExport instance
        manifest_export = ManifestExport(
            manifest=self.manifest,
            generated_by=self.user
        )
        
        # Save PDF file
        pdf_filename = f"MANIFEST_{self.manifest.manifest_number}.pdf"
        manifest_export.pdf_file.save(pdf_filename, ContentFile(pdf_content), save=False)
        
        # Save Excel file
        excel_filename = f"MANIFEST_{self.manifest.manifest_number}.xlsx"
        manifest_export.excel_file.save(excel_filename, ContentFile(excel_content), save=False)
        
        # Save the ManifestExport instance
        manifest_export.save()
    
    def finalize(self):

        
        # Wrap all operations in database transaction
        with transaction.atomic():
            # Validate manifest can be finalized
            self._validate_can_finalize()
            
            # Update manifest status
            self._update_manifest_status()
            
            # Update all bags
            self._update_bags()
            
            # Update all shipments
            self._update_shipments()
            
            # Create tracking events
            self._create_tracking_events()
            
            # Generate exports
            pdf_content, excel_content = self._generate_exports()
            
            # Store exports
            self._store_exports(pdf_content, excel_content)
        
        # Return export content
        return (pdf_content, excel_content)



class ManifestStatusUpdateService:
    """
    Handle manifest status updates with cascading changes.
    """
    
    def __init__(self, manifest, user):
        """
        Initialize service.
        
        Args:
            manifest: Manifest instance to update
            user: User performing the update
        """
        self.manifest = manifest
        self.user = user
    
    def update_to_departed(self):

        
        with transaction.atomic():
            # Update manifest status
            self.manifest.status = 'DEPARTED'
            self.manifest.save()
            
            # Update all bags to DISPATCHED
            for bag in self.manifest.bags.all():
                bag.status = 'DISPATCHED'
                bag.save()
            
            # Update shipments in bags to HANDED_TO_AIRLINE and create tracking events
            for bag in self.manifest.bags.all():
                for shipment in bag.shipment.all():
                    shipment.current_status = 'HANDED_TO_AIRLINE'
                    shipment.save()
                    
                    # Create tracking event
                    TrackingEvent.objects.create(
                        shipment=shipment,
                        status='HANDED_TO_AIRLINE',
                        description=f"Departed on flight {self.manifest.flight_number}",
                        location='Bangladesh Airport',
                        updated_by=self.user
                    )
            
            # Update individual shipments to HANDED_TO_AIRLINE and create tracking events
            for shipment in self.manifest.shipments.all():
                shipment.current_status = 'HANDED_TO_AIRLINE'
                shipment.save()
                
                # Create tracking event
                TrackingEvent.objects.create(
                    shipment=shipment,
                    status='HANDED_TO_AIRLINE',
                    description=f"Departed on flight {self.manifest.flight_number}",
                    location='Bangladesh Airport',
                    updated_by=self.user
                )
    
    def update_to_in_transit(self):

        from django.db import transaction
        from .models import TrackingEvent
        
        with transaction.atomic():
            # Update shipments in bags to IN_TRANSIT_TO_HK and create tracking events
            for bag in self.manifest.bags.all():
                for shipment in bag.shipment.all():
                    shipment.current_status = 'IN_TRANSIT_TO_HK'
                    shipment.save()
                    
                    # Create tracking event
                    TrackingEvent.objects.create(
                        shipment=shipment,
                        status='IN_TRANSIT_TO_HK',
                        description=f"In transit to Hong Kong on flight {self.manifest.flight_number}",
                        location='In Transit',
                        updated_by=self.user
                    )
            
            # Update individual shipments to IN_TRANSIT_TO_HK and create tracking events
            for shipment in self.manifest.shipments.all():
                shipment.current_status = 'IN_TRANSIT_TO_HK'
                shipment.save()
                
                # Create tracking event
                TrackingEvent.objects.create(
                    shipment=shipment,
                    status='IN_TRANSIT_TO_HK',
                    description=f"In transit to Hong Kong on flight {self.manifest.flight_number}",
                    location='In Transit',
                    updated_by=self.user
                )
def generate_invoice_pdf(shipment, shipper_name, shipper_address, 
                         consignee_name, consignee_address, line_items):
    """
    Generate commercial invoice PDF — EXACT match to the provided image.
    """
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    elements = []
    styles = getSampleStyleSheet()

    # ==================== PARAGRAPH STYLES ====================
    label_style = ParagraphStyle(
        'Label', parent=styles['Normal'], fontName='Helvetica-Bold', 
        fontSize=11, alignment=TA_CENTER, spaceBefore=3, spaceAfter=3
    )
    center_style = ParagraphStyle(
        'Center', parent=styles['Normal'], fontName='Helvetica', 
        fontSize=11, alignment=TA_CENTER, leading=14
    )
    left_style = ParagraphStyle(
        'Left', parent=styles['Normal'], fontName='Helvetica', 
        fontSize=11, alignment=TA_LEFT, leading=14
    )
    title_style = ParagraphStyle(
        'Title', parent=styles['Normal'], fontName='Helvetica-Bold', 
        fontSize=15, alignment=TA_CENTER
    )
    # Note: Preserving the "commerical" typo to remain a 1:1 match to the image
    sample_style = ParagraphStyle(
        'Sample', parent=styles['Normal'], fontName='Helvetica-Bold', 
        fontSize=12, alignment=TA_CENTER
    )
    awb_style = ParagraphStyle(
        'AWB', parent=styles['Normal'], fontName='Helvetica-Bold', 
        fontSize=20, alignment=TA_CENTER
    )

    elements.append(Spacer(1, 0.5*inch))

    # ==================== SHIPPER & CONSIGNEE ====================
    # Set up columns: [Shipper side, Empty Gap, Consignee side]
    sc_data = [
        [Paragraph('SHIPPER NAME', label_style), '', Paragraph('CONSIGNEE NAME', label_style)],
        [Paragraph(shipper_name.replace('\n', '<br/>'), center_style), '', Paragraph(consignee_name.replace('\n', '<br/>'), center_style)],
        ['', '', ''], # Spacer row
        [Paragraph('ADDRESS', label_style), '', Paragraph('ADDRESS', label_style)],
        [Paragraph(shipper_address.replace('\n', '<br/>'), left_style), '', Paragraph(consignee_address.replace('\n', '<br/>'), left_style)]
    ]

    sc_table = Table(sc_data, colWidths=[3.0*inch, 1.0*inch, 3.0*inch], rowHeights=[None, None, 0.2*inch, None, None])
    sc_table.setStyle(TableStyle([
        # Box *only* around the label rows (Row 0 and Row 3)
        ('BOX', (0, 0), (0, 0), 1.5, colors.black),
        ('BOX', (2, 0), (2, 0), 1.5, colors.black),
        ('BOX', (0, 3), (0, 3), 1.5, colors.black),
        ('BOX', (2, 3), (2, 3), 1.5, colors.black),
        
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(sc_table)
    elements.append(Spacer(1, 0.6*inch))

    # ==================== TITLE ====================
    elements.append(Paragraph('INVOICE and PACKING LIST', title_style))
    elements.append(Spacer(1, 0.3*inch))

    # ==================== PRODUCT TABLE ====================
    table_data = [
        ['', 'DESCRIPTION OF GOODS', 'WEIGHT\n(KG)', 'pcs', 'VALUE']
    ]

    total_value = 0
    for item in line_items:
        qty = float(item['quantity'])
        unit_v = float(item['unit_value'])
        line_total = qty * unit_v
        total_value += line_total

        table_data.append([
            '',
            str(item['description']).upper(),
            f"{float(item['weight']):g}KG",
            f"{int(qty)}PCS",
            f"{shipment.declared_currency}{line_total:.0f}"
        ])

    # Total Row setup: Place text in index 0 so it merges cleanly across the span
    table_data.append(['TOTAL AMOUNT', '', '', '', f"{shipment.declared_currency}{total_value:.0f}"])

    product_table = Table(table_data, colWidths=[0.4*inch, 3.2*inch, 1.0*inch, 1.0*inch, 1.4*inch])
    
    product_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1.5, colors.black),
        
        # Header formatting
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        
        # Body formatting
        ('ALIGN', (1, 1), (1, -2), 'LEFT'),    # Desc left aligned
        ('ALIGN', (2, 1), (-1, -2), 'CENTER'), # Weight, pcs, value centered
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        
        # Row height/padding
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),

        # Total Row Span and Formatting
        ('SPAN', (0, -1), (3, -1)),            # Merge columns 0 through 3 in the last row
        ('ALIGN', (0, -1), (0, -1), 'CENTER'), # Center 'TOTAL AMOUNT'
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), 
        ('ALIGN', (4, -1), (4, -1), 'CENTER'), # Center the final total value
        ('FONTSIZE', (0, -1), (-1, -1), 11),
    ]))

    elements.append(product_table)
    elements.append(Spacer(1, 0.6*inch))

    # ==================== FOOTER ====================
    elements.append(Paragraph('Sample of no commerical value', sample_style))
    elements.append(Spacer(1, 0.8*inch))
    elements.append(Paragraph(f'AWB# {shipment.awb_number}', awb_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer