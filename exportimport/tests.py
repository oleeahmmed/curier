from django.test import TestCase
from django.contrib.auth.models import User
from .models import Bag


class BagSaveMethodTestCase(TestCase):
    """Test the save method auto-generates bag_number"""
    
    def test_save_generates_bag_number_when_empty(self):
        """Test that save() generates bag_number when not set"""
        bag = Bag.objects.create(status='OPEN')
        
        # Verify bag_number was auto-generated
        self.assertIsNotNone(bag.bag_number)
        self.assertTrue(bag.bag_number.startswith('HDK-BAG-'))
        
    def test_save_preserves_existing_bag_number(self):
        """Test that save() doesn't change existing bag_number"""
        bag = Bag.objects.create(bag_number='HDK-BAG-999999', status='OPEN')
        original_number = bag.bag_number
        
        # Modify and save again
        bag.status = 'SEALED'
        bag.save()
        
        # Verify bag_number unchanged
        self.assertEqual(bag.bag_number, original_number)
        
    def test_first_bag_number_is_000001(self):
        """Test that first bag gets HDK-BAG-000001"""
        bag = Bag.objects.create(status='OPEN')
        
        self.assertEqual(bag.bag_number, 'HDK-BAG-000001')
        
    def test_sequential_bag_numbers(self):
        """Test that bag numbers are sequential"""
        bag1 = Bag.objects.create(status='OPEN')
        bag2 = Bag.objects.create(status='OPEN')
        bag3 = Bag.objects.create(status='OPEN')
        
        self.assertEqual(bag1.bag_number, 'HDK-BAG-000001')
        self.assertEqual(bag2.bag_number, 'HDK-BAG-000002')
        self.assertEqual(bag3.bag_number, 'HDK-BAG-000003')


class BagDeleteMethodTestCase(TestCase):
    """Test the delete method handles OPEN bags correctly"""
    
    def setUp(self):
        """Set up test data"""
        from .models import Shipment, TrackingEvent
        
        # Create a user for tracking events
        self.user = User.objects.create_user(username='testuser', password='testpass')
        
        # Create an OPEN bag
        self.open_bag = Bag.objects.create(bag_number='HDK-BAG-TEST001', status='OPEN')
        
        # Create shipments and add to bag
        self.shipment1 = Shipment.objects.create(
            awb_number='AWB001',
            current_status='BAGGED_FOR_EXPORT',
            direction='BD_TO_HK',
            shipper_name='Test Sender 1',
            shipper_phone='+8801234567890',
            shipper_address='123 Test St, Dhaka',
            recipient_name='Test Recipient 1',
            recipient_phone='+85212345678',
            recipient_address='456 Test Rd, Hong Kong',
            contents='Test items',
            declared_value=100.00,
            weight_estimated=2.5,
            payment_method='PREPAID'
        )
        self.shipment2 = Shipment.objects.create(
            awb_number='AWB002',
            current_status='BAGGED_FOR_EXPORT',
            direction='BD_TO_HK',
            shipper_name='Test Sender 2',
            shipper_phone='+8801234567891',
            shipper_address='789 Test Ave, Dhaka',
            recipient_name='Test Recipient 2',
            recipient_phone='+85212345679',
            recipient_address='789 Test Blvd, Hong Kong',
            contents='Test items 2',
            declared_value=150.00,
            weight_estimated=3.0,
            payment_method='PREPAID'
        )
        
        self.open_bag.shipment.add(self.shipment1, self.shipment2)
    
    def test_delete_open_bag_succeeds(self):
        """Test that deleting an OPEN bag succeeds"""
        from .models import Shipment
        
        bag_id = self.open_bag.id
        shipment1_id = self.shipment1.id
        shipment2_id = self.shipment2.id
        
        # Delete the bag
        self.open_bag.delete()
        
        # Verify bag is deleted
        self.assertFalse(Bag.objects.filter(id=bag_id).exists())
        
        # Verify shipments still exist and status reverted
        shipment1 = Shipment.objects.get(id=shipment1_id)
        shipment2 = Shipment.objects.get(id=shipment2_id)
        
        self.assertEqual(shipment1.current_status, 'RECEIVED_AT_BD')
        self.assertEqual(shipment2.current_status, 'RECEIVED_AT_BD')
    
    def test_delete_sealed_bag_raises_error(self):
        """Test that deleting a SEALED bag raises ValidationError"""
        from django.core.exceptions import ValidationError
        
        sealed_bag = Bag.objects.create(bag_number='HDK-BAG-TEST002', status='SEALED')
        
        with self.assertRaises(ValidationError) as context:
            sealed_bag.delete()
        
        self.assertIn('Cannot delete bag with status SEALED', str(context.exception))
    
    def test_delete_in_manifest_bag_raises_error(self):
        """Test that deleting an IN_MANIFEST bag raises ValidationError"""
        from django.core.exceptions import ValidationError
        
        manifest_bag = Bag.objects.create(bag_number='HDK-BAG-TEST003', status='IN_MANIFEST')
        
        with self.assertRaises(ValidationError) as context:
            manifest_bag.delete()
        
        self.assertIn('Cannot delete bag with status IN_MANIFEST', str(context.exception))
    
    def test_delete_dispatched_bag_raises_error(self):
        """Test that deleting a DISPATCHED bag raises ValidationError"""
        from django.core.exceptions import ValidationError
        
        dispatched_bag = Bag.objects.create(bag_number='HDK-BAG-TEST004', status='DISPATCHED')
        
        with self.assertRaises(ValidationError) as context:
            dispatched_bag.delete()
        
        self.assertIn('Cannot delete bag with status DISPATCHED', str(context.exception))
    
    def test_delete_creates_tracking_events(self):
        """Test that deleting a bag creates tracking events for shipments"""
        from .models import TrackingEvent
        
        # Count tracking events before deletion
        events_before = TrackingEvent.objects.count()
        
        # Delete the bag
        self.open_bag.delete()
        
        # Verify tracking events were created (2 shipments = 2 events)
        events_after = TrackingEvent.objects.count()
        self.assertEqual(events_after, events_before + 2)
        
        # Verify event details
        events = TrackingEvent.objects.filter(
            shipment__in=[self.shipment1.id, self.shipment2.id]
        ).order_by('id')
        
        for event in events:
            self.assertEqual(event.status, 'RECEIVED_AT_BD')
            self.assertIn('Removed from deleted bag HDK-BAG-TEST001', event.description)
            self.assertEqual(event.location, 'Bangladesh Warehouse')
            self.assertIsNone(event.updated_by)



class BagAdminDeleteTestCase(TestCase):
    """Test the BagAdmin delete_model method handles validation errors"""
    
    def setUp(self):
        """Set up test data"""
        from django.contrib.admin.sites import AdminSite
        from .admin import BagAdmin
        from django.test import RequestFactory
        from django.contrib.auth.models import User
        from django.contrib.messages.storage.fallback import FallbackStorage
        
        # Create admin user
        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@test.com',
            password='adminpass'
        )
        
        # Create BagAdmin instance
        self.site = AdminSite()
        self.bag_admin = BagAdmin(Bag, self.site)
        
        # Create request factory
        self.factory = RequestFactory()
        
    def test_delete_model_with_open_bag_succeeds(self):
        """Test that delete_model succeeds for OPEN bags"""
        from django.contrib.messages.storage.fallback import FallbackStorage
        
        # Create an OPEN bag
        bag = Bag.objects.create(bag_number='HDK-BAG-TEST100', status='OPEN')
        
        # Create a mock request
        request = self.factory.post('/admin/exportimport/bag/')
        request.user = self.admin_user
        
        # Add messages framework support
        setattr(request, 'session', 'session')
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        
        # Call delete_model
        self.bag_admin.delete_model(request, bag)
        
        # Verify bag is deleted
        self.assertFalse(Bag.objects.filter(bag_number='HDK-BAG-TEST100').exists())
        
        # Verify no error messages
        message_list = list(messages)
        error_messages = [m for m in message_list if m.level_tag == 'error']
        self.assertEqual(len(error_messages), 0)
    
    def test_delete_model_with_sealed_bag_shows_error(self):
        """Test that delete_model shows error message for SEALED bags"""
        from django.contrib.messages.storage.fallback import FallbackStorage
        from django.contrib import messages as django_messages
        
        # Create a SEALED bag
        bag = Bag.objects.create(bag_number='HDK-BAG-TEST101', status='SEALED')
        
        # Create a mock request
        request = self.factory.post('/admin/exportimport/bag/')
        request.user = self.admin_user
        
        # Add messages framework support
        setattr(request, 'session', 'session')
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        
        # Call delete_model
        self.bag_admin.delete_model(request, bag)
        
        # Verify bag still exists (not deleted)
        self.assertTrue(Bag.objects.filter(bag_number='HDK-BAG-TEST101').exists())
        
        # Verify error message was added
        message_list = list(messages)
        error_messages = [m for m in message_list if m.level_tag == 'error']
        self.assertEqual(len(error_messages), 1)
        self.assertIn('Cannot delete bag with status SEALED', str(error_messages[0]))
    
    def test_delete_model_with_in_manifest_bag_shows_error(self):
        """Test that delete_model shows error message for IN_MANIFEST bags"""
        from django.contrib.messages.storage.fallback import FallbackStorage
        
        # Create an IN_MANIFEST bag
        bag = Bag.objects.create(bag_number='HDK-BAG-TEST102', status='IN_MANIFEST')
        
        # Create a mock request
        request = self.factory.post('/admin/exportimport/bag/')
        request.user = self.admin_user
        
        # Add messages framework support
        setattr(request, 'session', 'session')
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        
        # Call delete_model
        self.bag_admin.delete_model(request, bag)
        
        # Verify bag still exists (not deleted)
        self.assertTrue(Bag.objects.filter(bag_number='HDK-BAG-TEST102').exists())
        
        # Verify error message was added
        message_list = list(messages)
        error_messages = [m for m in message_list if m.level_tag == 'error']
        self.assertEqual(len(error_messages), 1)
        self.assertIn('Cannot delete bag with status IN_MANIFEST', str(error_messages[0]))
    
    def test_delete_model_with_dispatched_bag_shows_error(self):
        """Test that delete_model shows error message for DISPATCHED bags"""
        from django.contrib.messages.storage.fallback import FallbackStorage
        
        # Create a DISPATCHED bag
        bag = Bag.objects.create(bag_number='HDK-BAG-TEST103', status='DISPATCHED')
        
        # Create a mock request
        request = self.factory.post('/admin/exportimport/bag/')
        request.user = self.admin_user
        
        # Add messages framework support
        setattr(request, 'session', 'session')
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        
        # Call delete_model
        self.bag_admin.delete_model(request, bag)
        
        # Verify bag still exists (not deleted)
        self.assertTrue(Bag.objects.filter(bag_number='HDK-BAG-TEST103').exists())
        
        # Verify error message was added
        message_list = list(messages)
        error_messages = [m for m in message_list if m.level_tag == 'error']
        self.assertEqual(len(error_messages), 1)
        self.assertIn('Cannot delete bag with status DISPATCHED', str(error_messages[0]))



class ShipmentRemoveViewTestCase(TestCase):
    """Test the ShipmentRemoveView functionality"""
    
    def setUp(self):
        """Set up test data"""
        from .models import Shipment, Manifest, TrackingEvent, StaffProfile, Location
        
        # Create a location
        self.location = Location.objects.create(
            name='Test Warehouse',
            location_type='WAREHOUSE',
            country='Bangladesh',
            city='Dhaka',
            address='123 Test St',
            phone='+8801234567890'
        )
        
        # Create a staff user
        self.staff_user = User.objects.create_user(
            username='staffuser',
            password='staffpass',
            is_staff=True
        )
        
        # Create staff profile
        self.staff_profile = StaffProfile.objects.create(
            user=self.staff_user,
            role='BD_STAFF',
            location=self.location,
            phone='+8801234567890',
            employee_id='EMP001',
            is_active=True
        )
        
        # Create a manifest in DRAFT status
        self.manifest = Manifest.objects.create(
            manifest_number='MF20260219001',
            flight_number='BG123',
            departure_date='2026-02-20',
            departure_time='10:00',
            status='DRAFT',
            created_by=self.staff_user
        )
        
        # Create a bag
        self.bag = Bag.objects.create(
            bag_number='HDK-BAG-TEST200',
            status='SEALED',
            weight=5.5
        )
        
        # Create shipments
        self.shipment1 = Shipment.objects.create(
            awb_number='DH2026021900001',
            current_status='BAGGED_FOR_EXPORT',
            direction='BD_TO_HK',
            shipper_name='Test Sender',
            shipper_phone='+8801234567890',
            shipper_address='123 Test St, Dhaka',
            recipient_name='Test Recipient',
            recipient_phone='+85212345678',
            recipient_address='456 Test Rd, Hong Kong',
            contents='Test items',
            declared_value=100.00,
            weight_estimated=2.5,
            payment_method='PREPAID'
        )
        
        # Add shipment to bag
        self.bag.shipment.add(self.shipment1)
        
        # Add bag to manifest
        self.manifest.bags.add(self.bag)
        self.manifest.calculate_totals()
    
    def test_remove_shipment_from_draft_manifest_succeeds(self):
        """Test that removing a shipment from a DRAFT manifest succeeds"""
        from django.test import Client
        import json
        
        client = Client()
        client.login(username='staffuser', password='staffpass')
        
        # Remove shipment
        response = client.post(
            f'/manifests/{self.manifest.id}/shipments/{self.shipment1.id}/remove/',
            content_type='application/json'
        )
        
        # Verify response
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.content)
        self.assertTrue(data['success'])
        self.assertIn('removed', data['message'].lower())
        
        # Verify shipment was removed from bag
        self.assertFalse(self.bag.shipment.filter(id=self.shipment1.id).exists())
        
        # Verify shipment status reverted to RECEIVED_AT_BD
        self.shipment1.refresh_from_db()
        self.assertEqual(self.shipment1.current_status, 'RECEIVED_AT_BD')
        
        # Verify tracking event was created
        from .models import TrackingEvent
        tracking_event = TrackingEvent.objects.filter(
            shipment=self.shipment1,
            status='RECEIVED_AT_BD'
        ).latest('timestamp')
        
        self.assertIn('Removed from bag', tracking_event.description)
        self.assertEqual(tracking_event.location, 'Bangladesh Warehouse')
        self.assertEqual(tracking_event.updated_by, self.staff_user)
    
    def test_remove_shipment_from_finalized_manifest_fails(self):
        """Test that removing a shipment from a FINALIZED manifest fails"""
        from django.test import Client
        import json
        
        # Finalize the manifest
        self.manifest.status = 'FINALIZED'
        self.manifest.save()
        
        client = Client()
        client.login(username='staffuser', password='staffpass')
        
        # Try to remove shipment
        response = client.post(
            f'/manifests/{self.manifest.id}/shipments/{self.shipment1.id}/remove/',
            content_type='application/json'
        )
        
        # Verify response
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertFalse(data['success'])
        self.assertIn('finalized', data['error'].lower())
        
        # Verify shipment was NOT removed from bag
        self.assertTrue(self.bag.shipment.filter(id=self.shipment1.id).exists())
    
    def test_remove_shipment_not_in_manifest_fails(self):
        """Test that removing a shipment not in the manifest fails"""
        from django.test import Client
        import json
        from .models import Shipment
        
        # Create another shipment not in the manifest
        other_shipment = Shipment.objects.create(
            awb_number='DH2026021900002',
            current_status='RECEIVED_AT_BD',
            direction='BD_TO_HK',
            shipper_name='Other Sender',
            shipper_phone='+8801234567891',
            shipper_address='789 Test Ave, Dhaka',
            recipient_name='Other Recipient',
            recipient_phone='+85212345679',
            recipient_address='789 Test Blvd, Hong Kong',
            contents='Other items',
            declared_value=150.00,
            weight_estimated=3.0,
            payment_method='PREPAID'
        )
        
        client = Client()
        client.login(username='staffuser', password='staffpass')
        
        # Try to remove shipment not in manifest
        response = client.post(
            f'/manifests/{self.manifest.id}/shipments/{other_shipment.id}/remove/',
            content_type='application/json'
        )
        
        # Verify response
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertFalse(data['success'])
        self.assertIn('not in this manifest', data['error'].lower())
    
    def test_remove_shipment_updates_bag_weight(self):
        """Test that removing a shipment updates the bag weight"""
        from django.test import Client
        
        # Record initial bag weight
        initial_weight = self.bag.weight
        shipment_weight = self.shipment1.weight_estimated
        
        client = Client()
        client.login(username='staffuser', password='staffpass')
        
        # Remove shipment
        response = client.post(
            f'/manifests/{self.manifest.id}/shipments/{self.shipment1.id}/remove/',
            content_type='application/json'
        )
        
        # Verify response
        self.assertEqual(response.status_code, 200)
        
        # Verify bag weight was updated
        self.bag.refresh_from_db()
        # Weight should be recalculated based on remaining shipments
        expected_weight = self.bag.calculate_total_weight()
        self.assertEqual(self.bag.weight, expected_weight)
    
    def test_remove_shipment_updates_manifest_totals(self):
        """Test that removing a shipment updates manifest totals"""
        from django.test import Client
        
        # Record initial totals
        initial_parcels = self.manifest.total_parcels
        
        client = Client()
        client.login(username='staffuser', password='staffpass')
        
        # Remove shipment
        response = client.post(
            f'/manifests/{self.manifest.id}/shipments/{self.shipment1.id}/remove/',
            content_type='application/json'
        )
        
        # Verify response
        self.assertEqual(response.status_code, 200)
        
        # Verify manifest totals were updated
        self.manifest.refresh_from_db()
        self.assertEqual(self.manifest.total_parcels, initial_parcels - 1)



class ManifestExcelGeneratorTestCase(TestCase):
    """Test the ManifestExcelGenerator service"""
    
    def setUp(self):
        """Set up test data"""
        from .models import Shipment, Manifest, Location, StaffProfile
        from datetime import date, time
        
        # Create a location
        self.location = Location.objects.create(
            name='Test Warehouse',
            location_type='WAREHOUSE',
            country='Bangladesh',
            city='Dhaka',
            address='123 Test St',
            phone='+8801234567890'
        )
        
        # Create a staff user
        self.staff_user = User.objects.create_user(
            username='staffuser',
            password='staffpass',
            is_staff=True
        )
        
        # Create staff profile
        self.staff_profile = StaffProfile.objects.create(
            user=self.staff_user,
            role='BD_STAFF',
            location=self.location,
            phone='+8801234567890',
            employee_id='EMP001',
            is_active=True
        )
        
        # Create a manifest
        self.manifest = Manifest.objects.create(
            manifest_number='MF20260219001',
            flight_number='BG123',
            departure_date=date(2026, 2, 20),
            departure_time=time(10, 0),
            mawb_number='MAWB123456',
            status='DRAFT',
            created_by=self.staff_user
        )
        
        # Create bags
        self.bag1 = Bag.objects.create(
            bag_number='HDK-BAG-TEST300',
            status='SEALED',
            weight=5.5
        )
        
        self.bag2 = Bag.objects.create(
            bag_number='HDK-BAG-TEST301',
            status='SEALED',
            weight=3.2
        )
        
        # Create shipments for bag1
        self.shipment1 = Shipment.objects.create(
            awb_number='DH2026021900001',
            current_status='BAGGED_FOR_EXPORT',
            direction='BD_TO_HK',
            shipper_name='Test Sender 1',
            shipper_country='Bangladesh',
            shipper_phone='+8801234567890',
            shipper_address='123 Test St, Dhaka',
            recipient_name='Test Recipient 1',
            recipient_country='Hong Kong',
            recipient_phone='+85212345678',
            recipient_address='456 Test Rd, Hong Kong',
            contents='Electronics',
            declared_value=100.00,
            declared_currency='USD',
            weight_estimated=2.5,
            payment_method='PREPAID',
            is_cod=False
        )
        
        self.shipment2 = Shipment.objects.create(
            awb_number='DH2026021900002',
            current_status='BAGGED_FOR_EXPORT',
            direction='BD_TO_HK',
            shipper_name='Test Sender 2',
            shipper_country='Bangladesh',
            shipper_phone='+8801234567891',
            shipper_address='789 Test Ave, Dhaka',
            recipient_name='Test Recipient 2',
            recipient_country='Hong Kong',
            recipient_phone='+85212345679',
            recipient_address='789 Test Blvd, Hong Kong',
            contents='Clothing',
            declared_value=50.00,
            declared_currency='USD',
            weight_estimated=3.0,
            payment_method='COD',
            is_cod=True
        )
        
        # Add shipments to bags
        self.bag1.shipment.add(self.shipment1)
        self.bag2.shipment.add(self.shipment2)
        
        # Add bags to manifest
        self.manifest.bags.add(self.bag1, self.bag2)
        self.manifest.calculate_totals()
    
    def test_excel_generator_creates_valid_excel(self):
        """Test that ManifestExcelGenerator creates a valid Excel file"""
        from .services import ManifestExcelGenerator
        from openpyxl import load_workbook
        from io import BytesIO
        
        # Generate Excel
        generator = ManifestExcelGenerator(self.manifest)
        excel_content = generator.generate()
        
        # Verify content is not empty
        self.assertIsNotNone(excel_content)
        self.assertGreater(len(excel_content), 0)
        
        # Load the Excel file to verify it's valid
        buffer = BytesIO(excel_content)
        workbook = load_workbook(buffer)
        worksheet = workbook.active
        
        # Verify header
        self.assertEqual(worksheet.cell(row=1, column=1).value, "CORNICHE INTERNATIONAL LTD")
        self.assertEqual(worksheet.cell(row=2, column=1).value, "MANIFEST")
        
        # Verify meta information
        self.assertEqual(worksheet.cell(row=4, column=1).value, "Agent Name:")
        self.assertEqual(worksheet.cell(row=4, column=2).value, "FAST LINE (DAC)")
        self.assertEqual(worksheet.cell(row=4, column=4).value, "Flight No.:")
        self.assertEqual(worksheet.cell(row=4, column=5).value, "BG123")
        
        self.assertEqual(worksheet.cell(row=5, column=1).value, "MAWB:")
        self.assertEqual(worksheet.cell(row=5, column=2).value, "MAWB123456")
        self.assertEqual(worksheet.cell(row=5, column=4).value, "Flight Date:")
        self.assertEqual(worksheet.cell(row=5, column=5).value, "20/02/2026")
        
        # Verify column headers (row 7)
        headers = ['No.', 'AWB No', 'Shipper', 'Consignee', 'Origin/Dest',
                   'PCS', 'Weight', 'Description', 'Value/Code', 'Remark', 'COD', 'Bag No']
        for col_idx, header in enumerate(headers, start=1):
            self.assertEqual(worksheet.cell(row=7, column=col_idx).value, header)
        
        # Verify shipment data (row 8 and 9)
        # First shipment
        self.assertEqual(worksheet.cell(row=8, column=1).value, 1)  # No.
        self.assertEqual(worksheet.cell(row=8, column=2).value, 'DH2026021900001')  # AWB No
        self.assertEqual(worksheet.cell(row=8, column=3).value, 'Test Sender 1')  # Shipper
        self.assertIn('Test Recipient 1', worksheet.cell(row=8, column=4).value)  # Consignee
        self.assertEqual(worksheet.cell(row=8, column=12).value, 'HDK-BAG-TEST300')  # Bag No
        
        # Second shipment
        self.assertEqual(worksheet.cell(row=9, column=1).value, 2)  # No.
        self.assertEqual(worksheet.cell(row=9, column=2).value, 'DH2026021900002')  # AWB No
        self.assertEqual(worksheet.cell(row=9, column=11).value, 'COD')  # COD indicator
        
        # Verify totals row exists
        totals_row = worksheet.cell(row=11, column=1).value
        self.assertIsNotNone(totals_row)
        self.assertIn('TOTAL SHIPMENTS', totals_row)
        self.assertIn('PCS', totals_row)
        self.assertIn('WEIGHT', totals_row)
    
    def test_excel_generator_handles_empty_manifest(self):
        """Test that ManifestExcelGenerator handles manifest with no bags"""
        from .services import ManifestExcelGenerator
        from openpyxl import load_workbook
        from io import BytesIO
        from datetime import date, time
        
        # Create empty manifest
        empty_manifest = Manifest.objects.create(
            manifest_number='MF20260219002',
            flight_number='BG456',
            departure_date=date(2026, 2, 21),
            departure_time=time(14, 0),
            status='DRAFT',
            created_by=self.staff_user
        )
        
        # Generate Excel
        generator = ManifestExcelGenerator(empty_manifest)
        excel_content = generator.generate()
        
        # Verify content is not empty
        self.assertIsNotNone(excel_content)
        self.assertGreater(len(excel_content), 0)
        
        # Load the Excel file
        buffer = BytesIO(excel_content)
        workbook = load_workbook(buffer)
        worksheet = workbook.active
        
        # Verify header and meta info still present
        self.assertEqual(worksheet.cell(row=1, column=1).value, "CORNICHE INTERNATIONAL LTD")
        self.assertEqual(worksheet.cell(row=2, column=1).value, "MANIFEST")
        
        # Verify totals show 0
        totals_row = worksheet.cell(row=9, column=1).value
        self.assertIn('TOTAL SHIPMENTS: 0', totals_row)
    
    def test_excel_generator_formats_consignee_multiline(self):
        """Test that consignee address is formatted with multiple lines"""
        from .services import ManifestExcelGenerator
        from openpyxl import load_workbook
        from io import BytesIO
        
        # Generate Excel
        generator = ManifestExcelGenerator(self.manifest)
        excel_content = generator.generate()
        
        # Load the Excel file
        buffer = BytesIO(excel_content)
        workbook = load_workbook(buffer)
        worksheet = workbook.active
        
        # Check consignee cell (row 8, column 4)
        consignee_value = worksheet.cell(row=8, column=4).value
        
        # Verify it contains both name and address
        self.assertIn('Test Recipient 1', consignee_value)
        self.assertIn('456 Test Rd, Hong Kong', consignee_value)
        
        # Verify it's multi-line (contains newline)
        self.assertIn('\n', consignee_value)



class ManifestFinalizationServiceTestCase(TestCase):
    """Test the ManifestFinalizationService"""
    
    def setUp(self):
        """Set up test data"""
        from .models import Shipment, Manifest, Location, StaffProfile
        from datetime import date, time
        
        # Create a location
        self.location = Location.objects.create(
            name='Test Warehouse',
            location_type='WAREHOUSE',
            country='Bangladesh',
            city='Dhaka',
            address='123 Test St',
            phone='+8801234567890'
        )
        
        # Create a staff user
        self.staff_user = User.objects.create_user(
            username='staffuser',
            password='staffpass',
            is_staff=True
        )
        
        # Create staff profile
        self.staff_profile = StaffProfile.objects.create(
            user=self.staff_user,
            role='BD_STAFF',
            location=self.location,
            phone='+8801234567890',
            employee_id='EMP001',
            is_active=True
        )
        
        # Create a manifest in DRAFT status
        self.manifest = Manifest.objects.create(
            manifest_number='MF20260219001',
            flight_number='BG123',
            departure_date=date(2026, 2, 20),
            departure_time=time(10, 0),
            mawb_number='MAWB123456',
            status='DRAFT',
            created_by=self.staff_user
        )
        
        # Create a bag
        self.bag = Bag.objects.create(
            bag_number='HDK-BAG-TEST400',
            status='SEALED',
            weight=5.5
        )
        
        # Create shipment
        self.shipment = Shipment.objects.create(
            awb_number='DH2026021900001',
            current_status='BAGGED_FOR_EXPORT',
            direction='BD_TO_HK',
            shipper_name='Test Sender',
            shipper_country='Bangladesh',
            shipper_phone='+8801234567890',
            shipper_address='123 Test St, Dhaka',
            recipient_name='Test Recipient',
            recipient_country='Hong Kong',
            recipient_phone='+85212345678',
            recipient_address='456 Test Rd, Hong Kong',
            contents='Test items',
            declared_value=100.00,
            declared_currency='USD',
            weight_estimated=2.5,
            payment_method='PREPAID'
        )
        
        # Add shipment to bag
        self.bag.shipment.add(self.shipment)
        
        # Add bag to manifest
        self.manifest.bags.add(self.bag)
        self.manifest.calculate_totals()
    
    def test_finalize_draft_manifest_succeeds(self):
        """Test that finalizing a DRAFT manifest succeeds"""
        from .services import ManifestFinalizationService
        
        # Create service and finalize
        service = ManifestFinalizationService(self.manifest, self.staff_user)
        pdf_content, excel_content = service.finalize()
        
        # Verify manifest status updated
        self.manifest.refresh_from_db()
        self.assertEqual(self.manifest.status, 'FINALIZED')
        self.assertEqual(self.manifest.finalized_by, self.staff_user)
        self.assertIsNotNone(self.manifest.finalized_at)
        
        # Verify bag status updated
        self.bag.refresh_from_db()
        self.assertEqual(self.bag.status, 'IN_MANIFEST')
        
        # Verify shipment status updated
        self.shipment.refresh_from_db()
        self.assertEqual(self.shipment.current_status, 'IN_EXPORT_MANIFEST')
        
        # Verify tracking event created
        from .models import TrackingEvent
        tracking_event = TrackingEvent.objects.filter(
            shipment=self.shipment,
            status='IN_EXPORT_MANIFEST'
        ).latest('timestamp')
        
        self.assertIn('Added to manifest', tracking_event.description)
        self.assertIn(self.manifest.manifest_number, tracking_event.description)
        self.assertEqual(tracking_event.location, 'Bangladesh Warehouse')
        self.assertEqual(tracking_event.updated_by, self.staff_user)
        
        # Verify exports generated
        self.assertIsNotNone(pdf_content)
        self.assertIsNotNone(excel_content)
        self.assertGreater(len(pdf_content), 0)
        self.assertGreater(len(excel_content), 0)
        
        # Verify ManifestExport created
        from .models import ManifestExport
        manifest_export = ManifestExport.objects.get(manifest=self.manifest)
        self.assertEqual(manifest_export.generated_by, self.staff_user)
        self.assertTrue(manifest_export.pdf_file)
        self.assertTrue(manifest_export.excel_file)
    
    def test_finalize_finalized_manifest_raises_error(self):
        """Test that finalizing a FINALIZED manifest raises ValidationError"""
        from .services import ManifestFinalizationService
        from django.core.exceptions import ValidationError
        
        # Finalize the manifest first
        self.manifest.status = 'FINALIZED'
        self.manifest.save()
        
        # Try to finalize again
        service = ManifestFinalizationService(self.manifest, self.staff_user)
        
        with self.assertRaises(ValidationError) as context:
            service.finalize()
        
        self.assertIn('Only DRAFT manifests can be finalized', str(context.exception))
    
    def test_finalize_empty_manifest_raises_error(self):
        """Test that finalizing a manifest with no bags raises ValidationError"""
        from .services import ManifestFinalizationService
        from django.core.exceptions import ValidationError
        from .models import Manifest
        from datetime import date, time
        
        # Create empty manifest
        empty_manifest = Manifest.objects.create(
            manifest_number='MF20260219002',
            flight_number='BG456',
            departure_date=date(2026, 2, 21),
            departure_time=time(14, 0),
            status='DRAFT',
            created_by=self.staff_user
        )
        
        # Try to finalize
        service = ManifestFinalizationService(empty_manifest, self.staff_user)
        
        with self.assertRaises(ValidationError) as context:
            service.finalize()
        
        self.assertIn('Manifest must have at least one bag', str(context.exception))
    
    def test_finalize_is_atomic(self):
        """Test that finalization is atomic - all or nothing"""
        from .services import ManifestFinalizationService
        from unittest.mock import patch
        
        # Mock _store_exports to raise an exception
        with patch.object(ManifestFinalizationService, '_store_exports', side_effect=Exception('Storage failed')):
            service = ManifestFinalizationService(self.manifest, self.staff_user)
            
            # Try to finalize - should fail
            with self.assertRaises(Exception):
                service.finalize()
        
        # Verify manifest status NOT changed (transaction rolled back)
        self.manifest.refresh_from_db()
        self.assertEqual(self.manifest.status, 'DRAFT')
        self.assertIsNone(self.manifest.finalized_by)
        self.assertIsNone(self.manifest.finalized_at)
        
        # Verify bag status NOT changed
        self.bag.refresh_from_db()
        self.assertEqual(self.bag.status, 'SEALED')
        
        # Verify shipment status NOT changed
        self.shipment.refresh_from_db()
        self.assertEqual(self.shipment.current_status, 'BAGGED_FOR_EXPORT')



class ManifestStatusUpdateServiceTestCase(TestCase):
    """Test the ManifestStatusUpdateService"""
    
    def setUp(self):
        """Set up test data"""
        from .models import Shipment, Manifest, Location, StaffProfile
        from datetime import date, time
        
        # Create a location
        self.location = Location.objects.create(
            name='Test Warehouse',
            location_type='WAREHOUSE',
            country='Bangladesh',
            city='Dhaka',
            address='123 Test St',
            phone='+8801234567890'
        )
        
        # Create a staff user
        self.staff_user = User.objects.create_user(
            username='staffuser',
            password='staffpass',
            is_staff=True
        )
        
        # Create staff profile
        self.staff_profile = StaffProfile.objects.create(
            user=self.staff_user,
            role='BD_STAFF',
            location=self.location,
            phone='+8801234567890',
            employee_id='EMP001',
            is_active=True
        )
        
        # Create a manifest in FINALIZED status
        self.manifest = Manifest.objects.create(
            manifest_number='MF20260219001',
            flight_number='BG123',
            departure_date=date(2026, 2, 20),
            departure_time=time(10, 0),
            mawb_number='MAWB123456',
            status='FINALIZED',
            created_by=self.staff_user,
            finalized_by=self.staff_user
        )
        
        # Create bags
        self.bag1 = Bag.objects.create(
            bag_number='HDK-BAG-TEST500',
            status='IN_MANIFEST',
            weight=5.5
        )
        
        self.bag2 = Bag.objects.create(
            bag_number='HDK-BAG-TEST501',
            status='IN_MANIFEST',
            weight=3.2
        )
        
        # Create shipments
        self.shipment1 = Shipment.objects.create(
            awb_number='DH2026021900001',
            current_status='IN_EXPORT_MANIFEST',
            direction='BD_TO_HK',
            shipper_name='Test Sender 1',
            shipper_country='Bangladesh',
            shipper_phone='+8801234567890',
            shipper_address='123 Test St, Dhaka',
            recipient_name='Test Recipient 1',
            recipient_country='Hong Kong',
            recipient_phone='+85212345678',
            recipient_address='456 Test Rd, Hong Kong',
            contents='Test items 1',
            declared_value=100.00,
            declared_currency='USD',
            weight_estimated=2.5,
            payment_method='PREPAID'
        )
        
        self.shipment2 = Shipment.objects.create(
            awb_number='DH2026021900002',
            current_status='IN_EXPORT_MANIFEST',
            direction='BD_TO_HK',
            shipper_name='Test Sender 2',
            shipper_country='Bangladesh',
            shipper_phone='+8801234567891',
            shipper_address='789 Test Ave, Dhaka',
            recipient_name='Test Recipient 2',
            recipient_country='Hong Kong',
            recipient_phone='+85212345679',
            recipient_address='789 Test Blvd, Hong Kong',
            contents='Test items 2',
            declared_value=150.00,
            declared_currency='USD',
            weight_estimated=3.0,
            payment_method='PREPAID'
        )
        
        # Add shipments to bags
        self.bag1.shipment.add(self.shipment1)
        self.bag2.shipment.add(self.shipment2)
        
        # Add bags to manifest
        self.manifest.bags.add(self.bag1, self.bag2)
        self.manifest.calculate_totals()
    
    def test_update_to_departed_succeeds(self):
        """Test that updating manifest to DEPARTED succeeds"""
        from .services import ManifestStatusUpdateService
        
        # Create service and update to departed
        service = ManifestStatusUpdateService(self.manifest, self.staff_user)
        service.update_to_departed()
        
        # Verify manifest status updated
        self.manifest.refresh_from_db()
        self.assertEqual(self.manifest.status, 'DEPARTED')
        
        # Verify all bags updated to DISPATCHED
        self.bag1.refresh_from_db()
        self.bag2.refresh_from_db()
        self.assertEqual(self.bag1.status, 'DISPATCHED')
        self.assertEqual(self.bag2.status, 'DISPATCHED')
        
        # Verify all shipments updated to HANDED_TO_AIRLINE
        self.shipment1.refresh_from_db()
        self.shipment2.refresh_from_db()
        self.assertEqual(self.shipment1.current_status, 'HANDED_TO_AIRLINE')
        self.assertEqual(self.shipment2.current_status, 'HANDED_TO_AIRLINE')
        
        # Verify tracking events created
        from .models import TrackingEvent
        
        # Check shipment1 tracking event
        tracking_event1 = TrackingEvent.objects.filter(
            shipment=self.shipment1,
            status='HANDED_TO_AIRLINE'
        ).latest('timestamp')
        
        self.assertIn('Departed on flight', tracking_event1.description)
        self.assertIn(self.manifest.flight_number, tracking_event1.description)
        self.assertEqual(tracking_event1.location, 'Bangladesh Airport')
        self.assertEqual(tracking_event1.updated_by, self.staff_user)
        
        # Check shipment2 tracking event
        tracking_event2 = TrackingEvent.objects.filter(
            shipment=self.shipment2,
            status='HANDED_TO_AIRLINE'
        ).latest('timestamp')
        
        self.assertIn('Departed on flight', tracking_event2.description)
        self.assertIn(self.manifest.flight_number, tracking_event2.description)
        self.assertEqual(tracking_event2.location, 'Bangladesh Airport')
        self.assertEqual(tracking_event2.updated_by, self.staff_user)
    
    def test_update_to_in_transit_succeeds(self):
        """Test that updating manifest to IN_TRANSIT succeeds"""
        from .services import ManifestStatusUpdateService
        
        # First update to DEPARTED
        service = ManifestStatusUpdateService(self.manifest, self.staff_user)
        service.update_to_departed()
        
        # Now update to IN_TRANSIT
        service.update_to_in_transit()
        
        # Verify all shipments updated to IN_TRANSIT_TO_HK
        self.shipment1.refresh_from_db()
        self.shipment2.refresh_from_db()
        self.assertEqual(self.shipment1.current_status, 'IN_TRANSIT_TO_HK')
        self.assertEqual(self.shipment2.current_status, 'IN_TRANSIT_TO_HK')
        
        # Verify tracking events created
        from .models import TrackingEvent
        
        # Check shipment1 tracking event
        tracking_event1 = TrackingEvent.objects.filter(
            shipment=self.shipment1,
            status='IN_TRANSIT_TO_HK'
        ).latest('timestamp')
        
        self.assertEqual(tracking_event1.description, 'In transit to Hong Kong')
        self.assertEqual(tracking_event1.location, 'In Transit')
        self.assertEqual(tracking_event1.updated_by, self.staff_user)
        
        # Check shipment2 tracking event
        tracking_event2 = TrackingEvent.objects.filter(
            shipment=self.shipment2,
            status='IN_TRANSIT_TO_HK'
        ).latest('timestamp')
        
        self.assertEqual(tracking_event2.description, 'In transit to Hong Kong')
        self.assertEqual(tracking_event2.location, 'In Transit')
        self.assertEqual(tracking_event2.updated_by, self.staff_user)
    
    def test_update_to_departed_is_atomic(self):
        """Test that update_to_departed is atomic - all or nothing"""
        from .services import ManifestStatusUpdateService
        from unittest.mock import patch
        from .models import TrackingEvent
        
        # Mock TrackingEvent.objects.create to raise an exception after first call
        original_create = TrackingEvent.objects.create
        call_count = [0]
        
        def mock_create(*args, **kwargs):
            call_count[0] += 1
            if call_count[0] > 1:
                raise Exception('Tracking event creation failed')
            return original_create(*args, **kwargs)
        
        with patch.object(TrackingEvent.objects, 'create', side_effect=mock_create):
            service = ManifestStatusUpdateService(self.manifest, self.staff_user)
            
            # Try to update - should fail
            with self.assertRaises(Exception):
                service.update_to_departed()
        
        # Verify manifest status NOT changed (transaction rolled back)
        self.manifest.refresh_from_db()
        self.assertEqual(self.manifest.status, 'FINALIZED')
        
        # Verify bag statuses NOT changed
        self.bag1.refresh_from_db()
        self.bag2.refresh_from_db()
        self.assertEqual(self.bag1.status, 'IN_MANIFEST')
        self.assertEqual(self.bag2.status, 'IN_MANIFEST')
        
        # Verify shipment statuses NOT changed
        self.shipment1.refresh_from_db()
        self.shipment2.refresh_from_db()
        self.assertEqual(self.shipment1.current_status, 'IN_EXPORT_MANIFEST')
        self.assertEqual(self.shipment2.current_status, 'IN_EXPORT_MANIFEST')
    
    def test_update_to_in_transit_is_atomic(self):
        """Test that update_to_in_transit is atomic - all or nothing"""
        from .services import ManifestStatusUpdateService
        from unittest.mock import patch
        from .models import TrackingEvent
        
        # First update to DEPARTED
        service = ManifestStatusUpdateService(self.manifest, self.staff_user)
        service.update_to_departed()
        
        # Mock TrackingEvent.objects.create to raise an exception
        with patch.object(TrackingEvent.objects, 'create', side_effect=Exception('Tracking event creation failed')):
            service = ManifestStatusUpdateService(self.manifest, self.staff_user)
            
            # Try to update - should fail
            with self.assertRaises(Exception):
                service.update_to_in_transit()
        
        # Verify shipment statuses NOT changed (transaction rolled back)
        self.shipment1.refresh_from_db()
        self.shipment2.refresh_from_db()
        self.assertEqual(self.shipment1.current_status, 'HANDED_TO_AIRLINE')
        self.assertEqual(self.shipment2.current_status, 'HANDED_TO_AIRLINE')



class ManifestDeleteViewTestCase(TestCase):
    """Test the ManifestDeleteView functionality"""
    
    def setUp(self):
        """Set up test data"""
        from .models import Shipment, Manifest, Location, StaffProfile
        from datetime import date, time
        
        # Create a location
        self.location = Location.objects.create(
            name='Test Warehouse',
            location_type='WAREHOUSE',
            country='Bangladesh',
            city='Dhaka',
            address='123 Test St',
            phone='+8801234567890'
        )
        
        # Create a staff user
        self.staff_user = User.objects.create_user(
            username='staffuser',
            password='staffpass',
            is_staff=True
        )
        
        # Create staff profile
        self.staff_profile = StaffProfile.objects.create(
            user=self.staff_user,
            role='BD_STAFF',
            location=self.location,
            phone='+8801234567890',
            employee_id='EMP001',
            is_active=True
        )
        
        # Create a manifest in DRAFT status
        self.draft_manifest = Manifest.objects.create(
            manifest_number='MF20260219001',
            flight_number='BG123',
            departure_date=date(2026, 2, 20),
            departure_time=time(10, 0),
            status='DRAFT',
            created_by=self.staff_user
        )
        
        # Create a bag
        self.bag = Bag.objects.create(
            bag_number='HDK-BAG-TEST600',
            status='SEALED',
            weight=5.5
        )
        
        # Create shipment
        self.shipment = Shipment.objects.create(
            awb_number='DH2026021900001',
            current_status='BAGGED_FOR_EXPORT',
            direction='BD_TO_HK',
            shipper_name='Test Sender',
            shipper_country='Bangladesh',
            shipper_phone='+8801234567890',
            shipper_address='123 Test St, Dhaka',
            recipient_name='Test Recipient',
            recipient_country='Hong Kong',
            recipient_phone='+85212345678',
            recipient_address='456 Test Rd, Hong Kong',
            contents='Test items',
            declared_value=100.00,
            declared_currency='USD',
            weight_estimated=2.5,
            payment_method='PREPAID'
        )
        
        # Add shipment to bag
        self.bag.shipment.add(self.shipment)
        
        # Add bag to manifest
        self.draft_manifest.bags.add(self.bag)
        self.draft_manifest.calculate_totals()
    
    def test_delete_draft_manifest_succeeds(self):
        """Test that deleting a DRAFT manifest succeeds (Requirement 7.1)"""
        from django.test import Client
        from .models import Manifest, Shipment
        import json
        
        client = Client()
        client.login(username='staffuser', password='staffpass')
        
        manifest_id = self.draft_manifest.id
        bag_id = self.bag.id
        shipment_id = self.shipment.id
        
        # Delete manifest
        response = client.post(
            f'/manifests/{manifest_id}/delete/',
            content_type='application/json'
        )
        
        # Verify response
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.content)
        self.assertTrue(data['success'])
        self.assertIn('deleted successfully', data['message'])
        
        # Verify manifest is deleted
        self.assertFalse(Manifest.objects.filter(id=manifest_id).exists())
        
        # Verify bags and shipments are NOT deleted (Requirements 7.5, 7.6)
        self.assertTrue(Bag.objects.filter(id=bag_id).exists())
        self.assertTrue(Shipment.objects.filter(id=shipment_id).exists())
    
    def test_delete_finalized_manifest_fails(self):
        """Test that deleting a FINALIZED manifest fails (Requirement 7.2)"""
        from django.test import Client
        from .models import Manifest
        import json
        from datetime import date, time
        
        # Create a finalized manifest
        finalized_manifest = Manifest.objects.create(
            manifest_number='MF20260219002',
            flight_number='BG456',
            departure_date=date(2026, 2, 21),
            departure_time=time(14, 0),
            status='FINALIZED',
            created_by=self.staff_user,
            finalized_by=self.staff_user
        )
        
        client = Client()
        client.login(username='staffuser', password='staffpass')
        
        # Try to delete finalized manifest
        response = client.post(
            f'/manifests/{finalized_manifest.id}/delete/',
            content_type='application/json'
        )
        
        # Verify response
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertFalse(data['success'])
        self.assertIn('Cannot delete finalized manifest', data['error'])
        
        # Verify manifest still exists
        self.assertTrue(Manifest.objects.filter(id=finalized_manifest.id).exists())
    
    def test_delete_departed_manifest_fails(self):
        """Test that deleting a DEPARTED manifest fails (Requirement 7.3)"""
        from django.test import Client
        from .models import Manifest
        import json
        from datetime import date, time
        
        # Create a departed manifest
        departed_manifest = Manifest.objects.create(
            manifest_number='MF20260219003',
            flight_number='BG789',
            departure_date=date(2026, 2, 22),
            departure_time=time(16, 0),
            status='DEPARTED',
            created_by=self.staff_user,
            finalized_by=self.staff_user
        )
        
        client = Client()
        client.login(username='staffuser', password='staffpass')
        
        # Try to delete departed manifest
        response = client.post(
            f'/manifests/{departed_manifest.id}/delete/',
            content_type='application/json'
        )
        
        # Verify response
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertFalse(data['success'])
        self.assertIn('Cannot delete departed manifest', data['error'])
        
        # Verify manifest still exists
        self.assertTrue(Manifest.objects.filter(id=departed_manifest.id).exists())
    
    def test_delete_arrived_manifest_fails(self):
        """Test that deleting an ARRIVED manifest fails (Requirement 7.4)"""
        from django.test import Client
        from .models import Manifest
        import json
        from datetime import date, time
        
        # Create an arrived manifest
        arrived_manifest = Manifest.objects.create(
            manifest_number='MF20260219004',
            flight_number='BG999',
            departure_date=date(2026, 2, 23),
            departure_time=time(18, 0),
            status='ARRIVED',
            created_by=self.staff_user,
            finalized_by=self.staff_user
        )
        
        client = Client()
        client.login(username='staffuser', password='staffpass')
        
        # Try to delete arrived manifest
        response = client.post(
            f'/manifests/{arrived_manifest.id}/delete/',
            content_type='application/json'
        )
        
        # Verify response
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertFalse(data['success'])
        self.assertIn('Cannot delete arrived manifest', data['error'])
        
        # Verify manifest still exists
        self.assertTrue(Manifest.objects.filter(id=arrived_manifest.id).exists())
    
    def test_delete_manifest_preserves_bags_and_shipments(self):
        """Test that deleting a manifest preserves bags and shipments (Requirements 7.5, 7.6)"""
        from django.test import Client
        from .models import Manifest, Shipment
        import json
        
        client = Client()
        client.login(username='staffuser', password='staffpass')
        
        # Record IDs before deletion
        manifest_id = self.draft_manifest.id
        bag_id = self.bag.id
        shipment_id = self.shipment.id
        
        # Delete manifest
        response = client.post(
            f'/manifests/{manifest_id}/delete/',
            content_type='application/json'
        )
        
        # Verify response
        self.assertEqual(response.status_code, 200)
        
        # Verify manifest is deleted
        self.assertFalse(Manifest.objects.filter(id=manifest_id).exists())
        
        # Verify bag still exists with same status
        bag = Bag.objects.get(id=bag_id)
        self.assertEqual(bag.bag_number, 'HDK-BAG-TEST600')
        self.assertEqual(bag.status, 'SEALED')
        
        # Verify shipment still exists with same status
        shipment = Shipment.objects.get(id=shipment_id)
        self.assertEqual(shipment.awb_number, 'DH2026021900001')
        self.assertEqual(shipment.current_status, 'BAGGED_FOR_EXPORT')
        
        # Verify shipment is still in bag
        self.assertTrue(bag.shipment.filter(id=shipment_id).exists())
